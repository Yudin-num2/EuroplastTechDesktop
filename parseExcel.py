import pandas as pd # type: ignore
from datetime import datetime
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def find_curr_file():
    curr_date = datetime.now().strftime("%Y-%m-%d")
    files = os.listdir('.')
    file_name = None
    for file in files:
        if file.startswith(curr_date) and file.endswith('.xlsx'):
            file_name = file
            return file_name
    if file_name is None:
        print("Файл не найден")
        return

def plot_histogram_to_excel(file_name):
    # file_name = find_curr_file()
    df = pd.read_excel(file_name)
    name_counts = df['Фамилия, имя'].value_counts().reset_index()
    name_counts.columns = ['Фамилия, имя', 'Количество']

    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet(title='Графики')
        for r_idx, row in name_counts.iterrows():
            worksheet.cell(row=r_idx + 1, column=1, value=row['Фамилия, имя'])
            worksheet.cell(row=r_idx + 2, column=2, value=row['Количество'])

        chart = BarChart()
        chart.show_legend = False
        chart.style = 10
        chart.title = "Количество выявленных"
        chart.x_axis.title = "Фамилия, имя"
        chart.y_axis.title = "Количество"
        max_row = len(name_counts)

        data = Reference(worksheet,
                         min_col=2,
                         min_row=1,
                         max_col=2,
                         max_row=max_row)
        categories = Reference(worksheet,
                               min_col=1,
                               min_row=1,
                               max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        worksheet.add_chart(chart, "E5")


# with pd.ExcelWriter(find_curr_file(), engine='openpyxl', mode='a') as writer:
#     workbook = writer.book
#     ws = workbook.create_sheet(title='Графики')
#     rows = [
#         ('Number', 'Batch 1', 'Batch 2'),
#         (2, 10, 30),
#         (3, 40, 60),
#         (4, 50, 70),
#         (5, 20, 10),
#         (6, 10, 40),
#         (7, 50, 30),
#     ]
#     for row in rows:
#         ws.append(row)


#     chart1 = BarChart()
#     chart1.style = 10
#     chart1.title = "Столбчатая диаграмма"
#     chart1.y_axis.title = 'Длина выборки'
#     chart1.y_axis.delete = False
#     chart1.x_axis.title = 'Номер теста'
#     chart1.x_axis.delete = False
#     data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=7)
#     categor = Reference(ws, min_col=1, min_row=2, max_row=7)
#     chart1.add_data(data, titles_from_data=True)
#     chart1.set_categories(categor)
#     ws.add_chart(chart1, "A10")


if __name__ == '__main__':
    plot_histogram_to_excel()
